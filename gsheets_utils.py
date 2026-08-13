import os
import re
import datetime
import win32com.client as win32
import pythoncom
import config_drive_paths
try:
    from judges_config import JERUSALEM_JUDGES, SUPREME_JUDGES
except ImportError:
    JERUSALEM_JUDGES = []
    SUPREME_JUDGES = []

def log(msg):
    # Temporary log function - mail_word_count_prod uses its own log, 
    # but we keep this for standalone gsheets_utils calls
    print(f"[gsheets_utils] {msg}")


# ============================================================================
# Excel-based district lookup  primary source of truth for district detection
# ============================================================================

# Excel paths (lazy-loaded once per process)
_JERUSALEM_EXCEL = r"C:\Users\yoel\OneDrive - Hever\טבלת מעקב בתי משפט 2023 מעודכן מתאריך 25.6.xlsx"
_JERUSALEM_SHEET = "2023"

def _get_south_excel():
    import config_drive_paths
    return config_drive_paths.scan_for_south_excel()

# Maps Jerusalem "סוג בית משפט" column values -> district key
_JERU_COURT_TO_DISTRICT = {
    "מחוזי":   "מחוזי_ירושלים",
    "שלום":    "שלום_ירושלים",
    "משפחה":   "משפחה_ירושלים",
    "תעבורה":  "משפחה_ירושלים",
    "עליון":   "ביהמש_העליון",
    "בית המשפט העליון": "ביהמש_העליון",
}

_SOUTH_COURT_TO_DISTRICT = {
    "מחוזי":       "מחוזי_באר_שבע",
    "שלום":        "שלום_באר_שבע",
    "אשקלון":      "שלום_אשקלון",
    "קריית גת":    "שלום_קריית_גת",
    "קרית גת":     "שלום_קריית_גת",
    "משפחה":       "שלום_באר_שבע",
    "אילת":         "שלום_באר_שבע",
}


def lookup_district_from_excel(case_num: str, date_obj=None) -> str | None:
    """
    Primary district detection: look up the case in Jerusalem and South Excel files.
    Returns a district key (e.g. 'שלום_ירושלים') or None if not found.

    Priority:
      1. Jerusalem Excel (most common for this system)
      2. South Excel
    """
    import pythoncom
    import win32com.client as win32

    if not case_num:
        return None

    def _try_excel(path, sheet_name, court_map, region_tag):
        if not path or not os.path.exists(path):
            return None
            
        try:
            import openpyxl
            # read_only=True bypasses Excel COM locks and doesn't load the file into memory fully
            # data_only=True reads calculated values instead of formulas
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            
            # Find sheet (handle case where sheet name might differ slightly)
            ws = None
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active  # fallback to first sheet
                
            if not ws: return None

            # Find header row
            h_row = 1
            # We must iterate rows in read_only mode
            rows_iter = ws.iter_rows(values_only=True)
            row_1 = next(rows_iter, None)
            row_2 = next(rows_iter, None)
            
            headers = [str(x).strip() if x else "" for x in (row_1 or [])]
            if not any(h in headers for h in ("מספר תיק", "תיק")):
                h_row = 2
                headers = [str(x).strip() if x else "" for x in (row_2 or [])]

            # Find key columns (0-indexed)
            idx_case  = next((i for i, h in enumerate(headers) if h in ("מספר תיק", "תיק")), None)
            idx_date  = next((i for i, h in enumerate(headers) if h == "תאריך"), None)
            idx_court = next((i for i, h in enumerate(headers) if h in (
                "סוג בית משפט", "בית משפט", "בית-משפט", "ביהמ\"ש", "סוג")), None)
            idx_judge = next((i for i, h in enumerate(headers) if h in (
                "שם השופט", "שופט", "שופט/ת", "שם שופט")), None)
            idx_city  = next((i for i, h in enumerate(headers) if h in (
                "עיר הקלטה", "עיר", "מקצוע")), None)
                
            if idx_case is None:
                wb.close()
                return None

            clean_case = case_num.strip().replace("/", "-").replace("_", "-")

            def _norm(s):
                parts = s.split("-")
                if len(parts) >= 3 and len(parts[-1]) == 2:
                    return "-".join(parts[:-1]) + "-20" + parts[-1]
                return s

            # Continue iterating rows
            for row in rows_iter:
                if len(row) <= idx_case: continue
                
                cell_case = str(row[idx_case] or "").strip().replace("/", "-").replace("_", "-")
                if not cell_case: continue
                
                if _norm(cell_case) != _norm(clean_case) and cell_case != clean_case:
                    continue

                # Case matches - check date if provided
                if date_obj and idx_date is not None and len(row) > idx_date:
                    cell_date = normalize_date(row[idx_date])
                    if cell_date and cell_date != normalize_date(date_obj):
                        continue  # Same case, different date -> keep searching

                # Found it
                judge_val = str(row[idx_judge] or "").strip() if idx_judge is not None and len(row) > idx_judge else None
                court_val = str(row[idx_court] or "").strip() if idx_court is not None and len(row) > idx_court else ""
                city_val  = str(row[idx_city] or "").strip() if idx_city is not None and len(row) > idx_city else ""
                
                res_district = None
                
                court_val_to_check = court_val if court_val else city_val
                
                if ("שלום" in court_val_to_check or not court_val_to_check) and "משפחה" in city_val:
                    res_district = "משפחה_ירושלים"
                else:
                    for key, district in court_map.items():
                        if key in court_val_to_check:
                            res_district = district
                            break
                            
                wb.close()
                
                if not res_district:
                    return {
                        "district": None,
                        "judge": judge_val,
                        "court_type": court_val,
                        "recording_city": city_val
                    }
                    
                return {
                    "district": res_district,
                    "judge": judge_val,
                    "court_type": court_val,
                    "recording_city": city_val
                }

            wb.close()
            return None
            
        except Exception as e:
            log(f"   [EXCEL-DISTRICT] Error reading {region_tag} Excel with openpyxl: {e}")
            return None

    # Try Jerusalem first
    result = _try_excel(_JERUSALEM_EXCEL, _JERUSALEM_SHEET, _JERU_COURT_TO_DISTRICT, "Jerusalem")
    if result:
        return result

    # Then South
    south_path = _get_south_excel()
    result = _try_excel(south_path, "שרות א באר שבע ודרום", _SOUTH_COURT_TO_DISTRICT, "South")
    return result


def normalize_date(val):
    if val is None: return None
    if isinstance(val, datetime.date) and not isinstance(val, datetime.datetime): return val
    if hasattr(val, 'date'): return val.date()
    if isinstance(val, datetime.datetime): return val.date()
    
    # Clean string
    s = str(val).strip().split(" ")[0]
    s = s.replace('\u200e', '').replace('\u200f', '').replace('-', '.').replace('/', '.')
    
    # --- Hebrew Month Support ---
    hebrew_months = {
        "ינואר": 1, "פברואר": 2, "מרץ": 3, "מרס": 3, "אפריל": 4, "מאי": 5, "יוני": 6,
        "יולי": 7, "אוגוסט": 8, "ספטמבר": 9, "אוקטובר": 10, "נובמבר": 11, "דצמבר": 12,
        "אפריל.": 4 # common trailing dot
    }
    
    # Check if string contains a Hebrew month name
    for m_name, m_num in hebrew_months.items():
        if m_name in s:
            # Try to find a day number before or after the month
            day_match = re.search(r'(\d{1,2})', s.replace(m_name, ''))
            if day_match:
                day = int(day_match.group(1))
                year = datetime.datetime.now().year # Default to current year
                # Look for a year like 2026 or 26
                year_match = re.search(r'(\d{4})|(\d{2})$', s)
                if year_match:
                    year_val = year_match.group(1) or year_match.group(2)
                    year = int(year_val)
                    if year < 100: year += 2000
                return datetime.date(year, m_num, day)

    # Try multiple numeric formats
    for fmt in [
        "%d.%m.%Y", "%d.%m.%y", "%Y.%m.%d", "%m.%d.%Y"
    ]:
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except: continue
        
    return None

def find_unique_hearing_row(ws, case_num, date_obj):
    """
    Enforces the 'one-to-one' rule: Matches both Case Number and Date.
    Returns row index if exactly one match found, else None/False.
    """
    if not case_num or not date_obj:
        log(f"[WARNING] Missing search criteria: Case={case_num}, Date={date_obj}")
        return None

    # Resolve date_obj to date only
    target_date = normalize_date(date_obj)
    clean_case = str(case_num).strip()

    # Determine Header Row (South usually 1, JLM usually 2)
    h_row = 1
    if "2023" in str(ws.Name) or "ירושלים" in str(ws.Parent.Name):
        # Heuristic for Jerusalem Excel
        if str(ws.Cells(2, 3).Value or "").strip() != "":
            h_row = 2

    from sys_process_utils import robust_com_call
    def _get_headers_and_last_row():
        last_col_val = ws.UsedRange.Columns.Count
        headers_val = [str(ws.Cells(h_row, c).Value or "").strip() for c in range(1, last_col_val + 1)]
        last_row_val = ws.UsedRange.Rows.Count
        return last_col_val, headers_val, last_row_val
        
    last_col, headers, last_row = robust_com_call(_get_headers_and_last_row, max_retries=5, delay=1.0)
    
    try:
        idx_case = headers.index("מספר תיק") + 1
        idx_date = headers.index("תאריך") + 1
    except ValueError:
        log(f"[ERROR] Headers 'מספר תיק' or 'תאריך' not found in {ws.Name}. Headers seen: {headers[:10]}")
        return None

    matches = []
    
    if last_row > h_row:
        def _get_raw_data():
            raw_cases_val = ws.Range(ws.Cells(h_row + 1, idx_case), ws.Cells(last_row, idx_case)).Value
            raw_dates_val = ws.Range(ws.Cells(h_row + 1, idx_date), ws.Cells(last_row, idx_date)).Value
            return raw_cases_val, raw_dates_val
            
        raw_cases, raw_dates = robust_com_call(_get_raw_data, max_retries=5, delay=1.0)
        
        if raw_cases and raw_dates:
            norm_clean = clean_case.replace('/', '-').replace('_', '-')
            for i in range(len(raw_cases) - 1, -1, -1):
                try: row_case_val = raw_cases[i][0]
                except: continue
                
                row_case = str(row_case_val or "").strip()
                norm_row = row_case.replace('/', '-').replace('_', '-')
                
                # Support combined case numbers by splitting both target and row cases
                target_parts = [x.strip() for x in re.split(r'[\s,;]+', norm_clean) if x.strip()]
                row_parts = [x.strip() for x in re.split(r'[\s,;]+', norm_row) if x.strip()]
                
                match_found = False
                for t_part in target_parts:
                    super_clean_target = re.sub(r'[^0-9]', '', t_part)
                    if not super_clean_target: continue
                    for r_part in row_parts:
                        super_clean_row = re.sub(r'[^0-9]', '', r_part)
                        if super_clean_target == super_clean_row:
                            match_found = True
                            break
                    if match_found:
                        break
                
                if match_found:
                    try: row_date_val = raw_dates[i][0]
                    except: continue
                    
                    row_date = normalize_date(row_date_val)
                    if row_date == target_date:
                        matches.append(h_row + 1 + i)
                        if len(matches) > 1:
                            log(f"[WARNING] [NON-UNIQUE] Multiple rows match {clean_case} on {target_date}: {matches}")
                            return "NON_UNIQUE"
    
    if len(matches) == 1:
        return matches[0]

    return None

def find_recording_region(case_num, date_obj=None):
    """
    Looks for a recording folder/file on F: drive for a given case.
    Returns: (region, found_date) or (None, None)
    """
    if not case_num:
        return None, None
        
    clean_target = case_num.strip()
    
    # 1. PRIORITY: Check Yoel's (Service B) folder first
    yoel_dir = r"F:\בתי משפט\דיונים ליואל"
    if os.path.exists(yoel_dir):
        try:
            for folder in os.listdir(yoel_dir):
                if clean_target in folder:
                    if date_obj:
                        date_str = f"{date_obj.day}_{date_obj.month}_{date_obj.year}"
                        if date_str in folder:
                            return "Yoel", date_obj
                    else:
                        match = re.search(r"_(\d{1,2})_(\d{1,2})_(\d{4})_", folder)
                        if match:
                            d, m, y = int(match.group(1)), int(match.group(2)), int(match.group(3))
                            return "Yoel", datetime.date(y, m, d)
        except: pass

    # 2. Check Jerusalem and South hierarchies (Scan current + 3 months back)
    today = datetime.date.today()
    scan_dates = [date_obj] if date_obj else [today]
    if not date_obj:
        curr = today.replace(day=1)
        for _ in range(3):
            curr = curr - datetime.timedelta(days=1)
            scan_dates.append(curr)
            curr = curr.replace(day=1)
        
    regions = ["Jerusalem", "South"]
    courts = ["מחוזי", "שלום", "עבודה", "משפחה", "תעבורה", "אחר"]

    for reg in regions:
        for sd in scan_dates:
            try:
                base_path = config_drive_paths.get_region_path(reg, sd)
                if not os.path.exists(base_path): continue
                for court in courts:
                    c_path = os.path.join(base_path, court)
                    if not os.path.exists(c_path): continue
                    date_folders = [d for d in os.listdir(c_path) if os.path.isdir(os.path.join(c_path, d))]
                    for df in date_folders:
                        parts = df.replace('-', '.').split('.')
                        if len(parts) >= 3:
                            try:
                                d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
                                if y < 100: y += 2000
                                folder_date = datetime.date(y, m, d)
                                if date_obj and folder_date != date_obj: continue
                                if any(clean_target in item for item in os.listdir(os.path.join(c_path, df))):
                                    return reg, folder_date
                            except: pass
            except: pass
    return None, None

def is_south_region(text, filename=""):
    south_keywords = ["באר שבע", "אילת", "אשדוד", "אשקלון", "קרית גת", "קריית גת", "דימונה", "ערד"]
    text_lower = text.lower()
    filename_lower = filename.lower()

    # Safety guard: if a Jerusalem or Supreme Court judge is present, do NOT flag as South.
    # A Jerusalem case can mention southern city names (party addresses, etc.) without being a South case.
    for judge in JERUSALEM_JUDGES + SUPREME_JUDGES:
        if judge in text_lower or judge in filename_lower:
            return False

    for kw in south_keywords:
        if kw in text_lower: return True
    for kw in south_keywords:
        if kw in filename_lower: return True
    if any(k in filename_lower for k in ["דרום", "ב\"ש", "באר-שבע", "אשקלון-"]): return True
    return False

def is_jerusalem_region(text, filename=""):
    """DETECTS if a case belongs to Jerusalem based on judge names or keywords."""
    text_lower = text.lower()
    filename_lower = filename.lower()

    # 1. Check Judge Names (Highest confidence)
    for judge in JERUSALEM_JUDGES + SUPREME_JUDGES:
        if judge in text_lower or judge in filename_lower:
            return True
            
    # 2. Check Jerusalem Specific Keywords
    jeru_keywords = ["ירושלים", "בית שמש", "ביתר", "מעלה אדומים", "עליון", "העליון", "בג\"ץ", "בגץ"]
    for kw in jeru_keywords:
        if kw in text_lower or kw in filename_lower:
            return True
            
    return False

def find_pending_hearings(case_num, transcriber_hint=None, lookback_days=14):
    """
    Looks for hearings for a case number in the last X days that have NO word count.
    Used for automated resolution of files missing a date in the filename.
    
    If multiple candidates exist, cross-references transcriber_hint with the Excel.
    """
    clean_case = str(case_num).strip().replace("/", "-").replace("_", "-")
    results = [] # List of (date, row_idx, region_tag, ws_obj)
    
    pythoncom.CoInitialize()
    excel = None
    try:
        excel = win32.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        
        today = datetime.date.today()
        start_date = today - datetime.timedelta(days=lookback_days)

        targets = [
            (_JERUSALEM_EXCEL, _JERUSALEM_SHEET, "Jerusalem", 2),
            (_get_south_excel(), "שרות א באר שבע ודרום", "South", 1)
        ]

        for path, sheet, tag, h_row in targets:
            if not path or not os.path.exists(path): continue
            try:
                wb = excel.Workbooks.Open(path, ReadOnly=True, UpdateLinks=0)
                ws = wb.Sheets(sheet)
                
                last_col = ws.UsedRange.Columns.Count
                headers = [str(ws.Cells(h_row, c).Value or "").strip() for c in range(1, last_col + 1)]
                
                idx_case = next((i+1 for i, h in enumerate(headers) if h in ("מספר תיק", "תיק")), None)
                idx_date = next((i+1 for i, h in enumerate(headers) if h == "תאריך"), None)
                idx_words = next((i+1 for i, h in enumerate(headers) if h in ("מילים", "מילים למתמלל", "מספר מילים")), None)
                idx_trans = next((i+1 for i, h in enumerate(headers) if h in ("מתמלל", "מתמלל/ת", "שם מתמלל")), None)
                
                if not idx_case or not idx_date:
                    wb.Close(False)
                    continue
                
                last_row = ws.UsedRange.Rows.Count
                for r in range(h_row + 1, last_row + 1):
                    # Check case
                    cell_case = str(ws.Cells(r, idx_case).Value or "").strip().replace("/", "-").replace("_", "-")
                    if not cell_case or clean_case not in cell_case: continue
                    
                    # Check date range (Today or Past)
                    row_date = normalize_date(ws.Cells(r, idx_date).Value)
                    if not row_date or row_date > today or row_date < start_date: continue
                    
                    # Check if already processed (words not empty)
                    words = ws.Cells(r, idx_words).Value if idx_words else None
                    if words and str(words).strip() not in ("", "0"): continue
                    
                    # If we have a match, collect it
                    trans_val = str(ws.Cells(r, idx_trans).Value or "").strip() if idx_trans else ""
                    results.append({
                        "date": row_date,
                        "row": r,
                        "region": tag,
                        "transcriber": trans_val,
                        "excel_path": path,
                        "sheet_name": sheet
                    })
                
                wb.Close(False)
            except Exception as e:
                log(f"   [Error reading {tag} Excel in find_pending]: {e}")
                
        # --- Logic for selection ---
        if not results: return None
        
        # If hint provided, try to filter by transcriber
        if transcriber_hint and len(results) > 1:
            hint_lower = transcriber_hint.lower()
            filtered = [r for r in results if hint_lower in r['transcriber'].lower() or r['transcriber'].lower() in hint_lower]
            if len(filtered) == 1:
                return filtered[0]
            if len(filtered) > 1:
                results = filtered # Narrowed but still ambiguous
        
        # If exactly one result remains (either naturally or after hint filter)
        if len(results) == 1:
            return results[0]
            
        return "AMBIGUOUS" # Multiple candidates found and couldn't resolve
        
    except Exception as e:
        log(f"   [DeepMatch Error]: {e}")
        return None
    finally:
        if excel:
            try: excel.Quit()
            except: pass
        pythoncom.CoUninitialize()

def unified_excel_update(case_num, date_obj, word_count, sent_timestamp=None, region="jerusalem"):
    """
    Atomic update for Excel:
    1. Updates word count.
    2. Updates 'Sent to Customer' timestamp.
    3. Saves and Closes.
    """
    if region.lower() == "south":
        path = _get_south_excel()
        sheet_name = "שרות א באר שבע ודרום"
        h_row = 1
    else:
        path = _JERUSALEM_EXCEL
        sheet_name = _JERUSALEM_SHEET
        h_row = 2

    if not path or not os.path.exists(path): return False

    pythoncom.CoInitialize()
    excel = None
    wb = None
    opened_by_us = False
    created_excel = False
    try:
        try:
            excel = win32.GetActiveObject("Excel.Application")
        except Exception:
            excel = win32.Dispatch("Excel.Application")
            created_excel = True
        excel.Visible = False
        excel.DisplayAlerts = False
        
        abs_path = os.path.abspath(path)
        for w in excel.Workbooks:
            try:
                if os.path.abspath(w.FullName).lower() == abs_path.lower():
                    wb = w
                    break
            except:
                pass
                
        # If it's already open but Excel is invisible, treat it as opened by us so we clean it up
        if wb and not excel.Visible:
            opened_by_us = True
                
        if not wb:
            def _open():
                return excel.Workbooks.Open(path)
            wb = robust_com_call(_open, max_retries=5, delay=2.0)
            opened_by_us = True
            
        if not wb:
            raise ValueError(f"Workbook could not be opened: {path}")
            
        ws = wb.Sheets(sheet_name)
        
        row_idx = find_unique_hearing_row(ws, case_num, date_obj)
        if not isinstance(row_idx, int):
            log(f"  [ERROR] unified_excel_update: Could not find unique row for {case_num} @ {date_obj}")
            if opened_by_us:
                wb.Close(False)
            return False
            
        last_col = ws.UsedRange.Columns.Count
        headers = [str(ws.Cells(h_row, c).Value or "").strip() for c in range(1, last_col + 1)]
        
        # Find Columns
        idx_words = next((i+1 for i, h in enumerate(headers) if h in ("מילים", "מילים למתמלל", "מספר מילים")), None)
        idx_sent = next((i+1 for i, h in enumerate(headers) if h in ("תאריך מסירה", "תאריך מסירה ללקוח", "נשלח ללקוח", "שליחה ללקוח", "מיילים")), None)
        
        # Handle Jerusalem specific 'מיילים' column if needed
        if region.lower() == "jerusalem" and not idx_sent:
             idx_sent = next((i+1 for i, h in enumerate(headers) if h == "מיילים"), None)

        if idx_words:
            ws.Cells(row_idx, idx_words).Value = word_count

        if idx_sent:
            ts = sent_timestamp or datetime.datetime.now().strftime("%d/%m/%Y")
            ws.Cells(row_idx, idx_sent).Value = ts

        def _save():
            wb.Save()
        robust_com_call(_save, max_retries=5, delay=2.0)

        # Verify the write actually persisted before declaring success
        if idx_words and word_count:
            verified = ws.Cells(row_idx, idx_words).Value
            if verified is None or str(verified).strip() in ("", "None", "0", "nan"):
                log(f"  [ERROR] Write verification FAILED for {case_num} row {row_idx} — cell still empty after save. Keeping file in Pending.")
                if opened_by_us:
                    wb.Close(False)
                return False
            log(f"  [SUCCESS] Verified word count {verified} in row {row_idx} for {case_num}")

        if opened_by_us:
            wb.Close(True)
        log(f"  [SUCCESS] Updated Word Count ({word_count}) and Sent status for {case_num} row {row_idx}")
        return True
    except Exception as e:
        log(f"  [CRITICAL] unified_excel_update failed: {e}")
        return False
    finally:
        if opened_by_us and wb:
            try: wb.Close(False)
            except: pass
        if created_excel and excel:
            try: excel.Quit()
            except: pass
        pythoncom.CoUninitialize()

from sys_process_utils import robust_com_call

def update_south_sheet(data_dict):
    """Legacy wrapper for update_south_sheet"""
    return unified_excel_update(
        data_dict.get('מספר תיק'), 
        data_dict.get('תאריך'), 
        data_dict.get('מספר מילים'), 
        region="south"
    )

def update_sent_dates(case_num, date_obj, sent_to_customer=True, sent_to_transcriber=False, region="jerusalem"):
    """Legacy wrapper for update_sent_dates"""
    return unified_excel_update(case_num, date_obj, None, region=region)

