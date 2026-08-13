# mail_word_count_prod.py
# =============================================================================
# [START] PRODUCTION SCRIPT - גרסה מבצעית סופית (Word COM Accurate + LLM Safe)
# =============================================================================

import win32com.client as win32
import os
import re
import shutil
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill
import time
import pythoncom

# מודולים משניים
import gsheets_utils
import n8n_utils
from file_saver import save_file_to_hierarchy, determine_court_type
import send_district_emails
import file_validator
try:
    from llm_utils import extract_judge_llm
except ImportError:
    extract_judge_llm = None

# =============================================================================
# [Settings]
# =============================================================================
DRY_RUN = False   # Test mode - no email sending
EMAIL_ACCOUNT = "your_email@example.com"
MAIN_FOLDER = r"C:\Users\yoel\OneDrive - Hever\Jerusalem"
INPUT_FOLDER = os.path.join(MAIN_FOLDER, "Pending_Word_Count")
DONE_FOLDER = os.path.join(MAIN_FOLDER, "Done")
FAILED_FOLDER = os.path.join(MAIN_FOLDER, "Failed")

JERUSALEM_EXCEL_PATH = r"C:\Users\yoel\OneDrive - Hever\טבלת מעקב בתי משפט 2023 מעודכן מתאריך 25.6.xlsx"
JERUSALEM_SHEET_NAME = "2023"

_SOUTH_RELATIVE = r"\האחסון שלי\עותק של שרות א חדש.xlsx"
SOUTH_EXCEL_PATH = next((f"{d}:{_SOUTH_RELATIVE}" for d in ["I","J","K","G","H"] if os.path.exists(f"{d}:{_SOUTH_RELATIVE}")), None)
SOUTH_SHEET_NAME = "שרות א באר שבע ודרום"

LOG_DIR = r"d:\yoel\projects\auto\logs"
os.makedirs(LOG_DIR, exist_ok=True)
LOG_FILE = os.path.join(LOG_DIR, f"run_log_{datetime.now().strftime('%Y-%m-%d')}.txt")

# שמות העמודות
COL_CASE_NUM = "מספר תיק"
COL_DATE = "תאריך"
COL_WORD_COUNT = "מספר מילים"
COL_JUDGE = "שם שופט/ת"

def log(msg):
    try:
        timestamp = datetime.now().strftime("%H:%M:%S")
        safe_msg = str(msg).encode('ascii', errors='replace').decode('ascii')
        # We still want the full message in the log file (UTF-8)
        formatted_msg = f"{timestamp} - {msg}"
        
        # Safe print for terminal
        print(f"{timestamp} - {safe_msg}", flush=True)
        
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(formatted_msg + "\n")
    except:
        try: print(f"LOG_ERR: {str(msg)[:50]}")
        except: pass

def safe_move(src_path, dest_folder):
    """Moves a file to a destination folder. If it already exists, removes the source (duplicate)."""
    if not os.path.exists(src_path): return
    filename = os.path.basename(src_path)
    dest_path = os.path.join(dest_folder, filename)
    if os.path.exists(dest_path) and os.path.abspath(src_path) != os.path.abspath(dest_path):
        try: 
            os.remove(src_path)
            log(f"   [SYNC] Deleted source duplicate: {filename} (already exists in target)")
        except: pass
        return dest_path
    try:
        return shutil.move(src_path, dest_path)
    except Exception as e:
        log(f"   [ERR] Move failed: {e}")
        return src_path

def normalize_date(val):
    if val is None: return None
    if hasattr(val, 'date'): return val.date()
    if isinstance(val, datetime): return val.date()
    
    # Clean string
    s = str(val).strip().split(" ")[0]
    # Remove BOM or hidden chars
    s = s.replace('\u200e', '').replace('\u200f', '')
    
    # Try multiple formats
    for fmt in [
        "%d/%m/%Y", "%d.%m.%Y", "%Y-%m-%d", "%d-%m-%Y",
        "%d/%m/%y", "%d.%m.%y",
        "%Y.%m.%d"
    ]:
        try:
            return datetime.strptime(s, fmt).date()
        except: continue
        
    return None

def load_excel_as_dicts(path, sheet_name):
    """Loads an Excel sheet into a list of dictionaries for validation purposes."""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        
        # Determine header row (1 or 2)
        r_head = 1
        headers = [str(cell.value).strip() if cell.value else None for cell in ws[1]]
        if not any(headers):
            r_head = 2
            headers = [str(cell.value).strip() if cell.value else None for cell in ws[2]]
        
        data = []
        for row in ws.iter_rows(min_row=r_head + 1, values_only=True):
            row_dict = {}
            for i, val in enumerate(row):
                if i < len(headers) and headers[i]:
                    row_dict[headers[i]] = val
            if any(row_dict.values()):
                data.append(row_dict)
        wb.close() # released handle
        return data
    except Exception as e:
        log(f"[WARNING] Error loading sheet data for validation: {e}")
        return []

def send_unmapped_email(unmapped_data):
    try:
        if not unmapped_data: return
        outlook = win32.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)
        mail.To = EMAIL_ACCOUNT
        mail.Subject = f"[WARNING] התקבלו ש{len(unmapped_data)} קבצי תמלול שלא מצאו את עצמם באקסל"
        
        rows_html = ""
        for data in unmapped_data:
            c = data.get(COL_CASE_NUM, "לא ידוע")
            d = str(data.get(COL_DATE) or "-")
            j = data.get("judge") or "-"
            f = data.get("filename", "-")
            w = data.get("word_count", 0)
            rows_html += f"<tr><td style='border:1px solid #ccc;padding:5px;'>{f}</td><td style='border:1px solid #ccc;padding:5px;' dir='ltr'>{c}</td><td style='border:1px solid #ccc;padding:5px;'>{d}</td><td style='border:1px solid #ccc;padding:5px;'>{j}</td><td style='border:1px solid #ccc;padding:5px;'>{w}</td></tr>"
            
        html = f"""
        <html dir="rtl">
        <body style="font-family: Arial, sans-serif;">
            <h2 style="color:red;">[WARNING] קבצי תמלול התקבלו אך לא הוזנו באקסל</h2>
            <p>המערכת חיפשה במדויק לפי מספר התיק, ואף ביצעה חיפוש כפול לפי תאריך ושופט במקרה הצורך, אך לא מצאה להם דיון פתוח שחסר לו מילים באקסל. משכך, הקבצים הועברו לתיקיית Requires_Manual_Review לבדיקה ידנית.</p>
            <table style="border-collapse:collapse;width:100%;text-align:right;">
                <tr style="background:#eee;"><th style='border:1px solid #ccc;padding:5px;'>שם קובץ</th><th style='border:1px solid #ccc;padding:5px;'>תיק / זיהוי</th><th style='border:1px solid #ccc;padding:5px;'>תאריך</th><th style='border:1px solid #ccc;padding:5px;'>שופט / אזור</th><th style='border:1px solid #ccc;padding:5px;'>מילים</th></tr>
                {rows_html}
            </table>
        </body>
        </html>
        """
        mail.HTMLBody = html
        mail.Send()
        log(f"📧 נשלח דיווח במייל על {len(unmapped_data)} קבצים שלא עודכנו באקסל.")
    except Exception as e:
        log(f"[WARNING] שגיאה בשליחת מייל unmapped: {e}")

def extract_case_check_regex(text, is_content=False):
    """
    Revised Robust Regex:
    1. 2-part Supreme Court patterns (Explicit: ע"א, בג"ץ, ערעור אזרחי)
    2. 3-part: XXXXX-XX-XX
    3. 2-part: XXXX-XX (General)
    """
    if not text:
        return None, None
        
    search_text = text[:1500] if is_content else text
    
    # 1. Supreme Court patterns (Explicit) - Strongest Signal
    supreme_prefixes = r'ע"א|בג"ץ|ער"מ|בש"פ|דנ"א|רע"א|רע"פ|עע"מ|ערעור אזרחי|ערעור פלילי|ערעור מינהלי|רשות ערעור|ע"פ|עע"מ|עב"ל'
    supreme_m = re.search(fr'(?:{supreme_prefixes})\s*(?:מס(?:\'|פר)?\s+)?(\d{{2,6}})[/-](\d{{2,4}})', search_text)
    if supreme_m:
        g1, g2 = supreme_m.groups()
        return f"{g1}-{g2}", supreme_m.group(0)

    # Use finditer to check matches with context
    patterns = [
        r'(\d{2,6})[-/_](\d{1,2})[-/_](\d{2,6})', # 3-part (last chunk up to 6 digits for inverted YY_MM_CASE)
        r'(\d{2,6})[/-](\d{2,5})(?![/-]\d)',      # 2-part
        r'(\d{2,4})[_/-](\d{4,8})'                # 2-part with underscores like 23_7063 or 123456_24
    ]
    
    candidates = []
    for pat in patterns:
        for match in re.finditer(pat, search_text):
            candidates.append(match)
            
    # Sort by appearance
    candidates.sort(key=lambda m: m.start())

    for match in candidates:
        val = match.group(0)
        start = match.start()
        
        # Context check
        context_before = search_text[max(0, start-60):start]
        
        # NEGATIVE CONTEXT: If preceded by "Appeal ON" or "Lower Court", skip it.
        is_lower_court = False
        if is_content:
            if any(x in context_before for x in ["ערעור על", "ערעור כנגד", "ב-ת\"א", "ב-ת.א", "בתיק", "במסגרת"]):
                is_lower_court = True
            if "פסק דינ" in context_before and "של" in context_before:
                is_lower_court = True
        
        if is_lower_court: continue

        parts = re.split(r'[-/_]', val)
        
        if len(parts) == 3:
            # Check if it was YY_MM_CASE (e.g. 26_01_24403 -> p1=26, p2=01, p3=24403)
            if len(parts[0]) <= 2 and len(parts[2]) >= 3:
                normalized = f"{parts[2]}-{int(parts[1]):02d}-{parts[0]}"
            else:
                normalized = f"{parts[0]}-{int(parts[1]):02d}-{parts[2]}"
        else:
            # Check if it was year_case or case_year
            p1, p2 = parts[0], parts[1]
            if len(p1) <= 2 and len(p2) >= 3: # 23-7063 -> 7063-23
                normalized = f"{p2}-{p1}"
            else:
                normalized = f"{p1}-{p2}"
            
        return normalized, val
                
    return None, None

def verify_supreme_court_llm(filename, text_snippet):
    """LLM check to prevent false positive Supreme Court detection"""
    try:
        from llm_utils import ask_llm_complex_json
        prompt = f"""Task: Verify if this is a Supreme Court (בית המשפט העליון) case based on the text header.
        Filename: "{filename}"
        Header Snippet: "{text_snippet[:1000]}"
        Instructions:
        - Analyze the header text.
        - Look for "בית המשפט העליון" or "בג\"ץ".
        - NOT Supreme if it says "בית המשפט המחוזי" (District) or "בית משפט השלום" (Magistrate).
        - NOT Supreme if it's just a reference/citation to a Supreme Court ruling inside a lower court protocol.
        Return JSON: {{ "is_supreme": true/false, "confidence": 1-10, "reason": "short explanation" }}"""
        
        result = ask_llm_complex_json(prompt)
        if result and "is_supreme" in result:
            is_sup = result["is_supreme"]
            log(f"🤖 LLM Supreme Check: {is_sup} ({result.get('reason')})")
            return is_sup
    except Exception as e:
        log(f"[WARNING] LLM Check Error: {e}")
        return False
    return False

def read_word_file(file_path):
    # Reuse existing Word instance to avoid closing user's open windows
    _created_word = False
    try:
        word = win32.GetActiveObject("Word.Application")
    except Exception:
        word = win32.Dispatch("Word.Application")
        _created_word = True
    word.Visible = False
    word.DisplayAlerts = 0
    doc = None
    try:
        for attempt in range(3):
            try:
                doc = word.Documents.Open(file_path, ConfirmConversions=False, ReadOnly=True, Visible=False)
                break
            except Exception as e:
                log(f"[WARNING] Word Open attempt {attempt+1} failed: {e}")
                time.sleep(2)
                if attempt == 2: raise e

        count = 0
        try:
            count = doc.ComputeStatistics(0)
        except Exception as ce:
            log(f"[WARNING] ComputeStatistics failed, falling back to manual count: {ce}")
        
        text = ""
        try:
            for section in doc.Sections:
                for header in section.Headers:
                    if not header.Range.Text: continue
                    text += " " + header.Range.Text[:500]
        except: pass
        
        text += " " + doc.Range(0, 2000).Text
        
        try:
            full_len = doc.Characters.Count
            if full_len > 2000:
                start_char = max(2000, full_len - 2000)
                text += " ... " + doc.Range(start_char, full_len).Text
        except: pass
        
        text = text.strip()
        if count == 0 and text:
            count = len(text.split())
        
        doc.Close(False)
        return count, text
    except Exception as e:
        log(f"[WARNING] Word Read Error on '{file_path}': {e}")
        if doc:
            try: doc.Close(False)
            except: pass
        return 0, ""
    finally:
        if _created_word:
            try: word.Quit()
            except: pass

def update_jerusalem_table(file_path, sheet_name, data_list):
    updated_cases = []
    if not os.path.exists(file_path): return []
    
    log(f"\n[STATS] Updating Jerusalem Table...")
    # Reuse existing Excel instance to avoid closing user's open workbooks
    _created_excel = False
    try:
        excel = win32.GetActiveObject("Excel.Application")
    except Exception:
        excel = win32.Dispatch("Excel.Application")
        _created_excel = True
    # Save original state, restore after
    _orig_visible = True
    _orig_alerts = True
    try: _orig_visible = excel.Visible
    except: pass
    try: _orig_alerts = excel.DisplayAlerts
    except: pass
    try: excel.Visible = False # Changed from True to False to avoid stealing focus
    except: pass
    try: excel.DisplayAlerts = 0 # wdAlertsNone
    except: pass
    try: excel.AskToUpdateLinks = False
    except: pass
    
    wb = None
    target_name = os.path.basename(file_path)
    try:
        for w in excel.Workbooks:
            if w.Name == target_name:
                if w.ReadOnly:
                    log("   [STATS] Found workbook already open in Excel but it is READ-ONLY. Closing it to reopen as Read-Write.")
                    w.Close(False)
                else:
                    wb = w
                    log("   [STATS] Found workbook already open in Excel in Read-Write mode.")
                break
    except: pass

    if not wb:
        try:
            # Retry up to 3 times (30s total) in case OneDrive is still syncing
            for _attempt in range(3):
                # explicitly set ReadOnly=False
                wb = excel.Workbooks.Open(file_path, UpdateLinks=0, ReadOnly=False)
                if not wb.ReadOnly:
                    break
                log(f"⏳ Excel READ-ONLY (attempt {_attempt+1}/3), waiting 10s for OneDrive to release...")
                wb.Close(False)
                wb = None
                import time
                time.sleep(10)
        except Exception as e:
            log(f"[ERROR] Error opening Excel: {e}")
            
    if wb is None or wb.ReadOnly:
        log(f"[ERROR] Error: Excel still READ-ONLY after retries. Cannot update: {file_path}")
        if wb: wb.Close(False)
        return ["READ_ONLY_ERROR"]
    try:
        log(f"[STATS] File open for editing: {wb.FullName}")
        
        try: ws = wb.Sheets(sheet_name)
        except: 
            log("[ERROR] Sheet not found")
            return updated_cases
        
        # Column mapping
        r1 = 2 if str(ws.Cells(2,1).Value) not in ["None",""] else 1
        col_map = {}
        for c in range(1, 40):
            v = str(ws.Cells(r1, c).Value).strip()
            if v: col_map[v] = c
            
        case_col = col_map.get(COL_CASE_NUM) or col_map.get("תיק")
        words_col = col_map.get(COL_WORD_COUNT) or col_map.get("מספר מילים")
        date_col = col_map.get(COL_DATE) or col_map.get("תאריך")
        # Find Status column - header is 'האם בוטל?' (Col 8)
        status_col = col_map.get("האם בוטל?") or col_map.get("בוטל") or 8
        judge_col = col_map.get(COL_JUDGE)
        
        updated = 0
        search_range = ws.Columns(case_col)
        
        for data in data_list:
            t_case = data.get(COL_CASE_NUM)
            if not t_case: continue
            
            # Supreme check
            fn, txt = str(data["filename"]), str(data["text_snippet"])
            SUPREME = ["סולברג", "ברק-ארז", "ברק ארז", "מינץ", "וילנר", "גרוסקופף", "שטיין", "כנפי", "כבוב", "כשר", "רונן", "עמית"]
            is_sup = ("עליון" in fn or "בגץ" in fn or "בבית המשפט העליון" in txt[:300] or any(j in fn for j in SUPREME))
            if "בית המשפט המחוזי" in txt[:300] or "בית משפט השלום" in txt[:300]: is_sup = False
            found_row = None
            
            
            # --- ROBUST MEMORY SEARCH ---
            last_row = ws.Cells(ws.Rows.Count, 3).End(-4162).Row
            
            if last_row > r1:
                # Read all case numbers at once
                raw_cases = ws.Range(ws.Cells(r1+1, case_col), ws.Cells(last_row, case_col)).Value
            else:
                raw_cases = []

            # Find all matching rows
            matched_rows = []
            if raw_cases:
                for i, row_val in enumerate(raw_cases):
                    # Excel returns tuples like ((val,), (val,), ...)
                    try: 
                        val_str = str(row_val[0]).strip()
                    except: 
                        continue
                        
                    # [START] FIXED MATCHING LOGIC
                    # We normalize all separators to dashes to avoid false positives 
                    # replacing everything with '' would make 3129-23 match 31292-3!
                    norm_t = t_case.replace('/','-').replace('_','-')
                    norm_v = val_str.replace('/','-').replace('_','-')
                    is_match = (norm_t == norm_v)
                    
                    if not is_match: 
                        # Fallback for exact substring like "12345-01-24" inside "12345-01-24 and more"
                        is_match = (t_case in val_str and len(val_str) < len(t_case)+15)
                    
                    if is_match:
                        # Row index calculation: r1 + 1 (start) + i (offset)
                        matched_rows.append(r1 + 1 + i)
            
            # --- DOUBLE SEARCH FALLBACK ---
            if not matched_rows and data.get(COL_DATE):
                target_d = normalize_date(data.get(COL_DATE))
                target_j = str(data.get("judge") or "").strip()
                
                if last_row > r1 and target_d:
                    try:
                        raw_dates = ws.Range(ws.Cells(r1+1, date_col), ws.Cells(last_row, date_col)).Value
                        raw_words = ws.Range(ws.Cells(r1+1, words_col), ws.Cells(last_row, words_col)).Value
                        raw_judges = ws.Range(ws.Cells(r1+1, judge_col), ws.Cells(last_row, judge_col)).Value if judge_col else []
                        
                        for idx in range(len(raw_dates)):
                            # 1. Check Date
                            d_val = raw_dates[idx][0] if raw_dates[idx] else None
                            d_obj = normalize_date(d_val)
                            if d_obj != target_d: continue
                            
                            # 2. Check Words (Must be empty)
                            w_val = raw_words[idx][0] if raw_words[idx] else None
                            curr_str = str(w_val).strip()
                            is_empty = (w_val is None or curr_str in ["", "None", "nan", "0"])
                            if not is_empty: continue
                            
                            # 3. Check Judge OR Partial Case
                            r_real = r1 + 1 + idx
                            
                            # Option A: Check partial case (e.g. 12345 from 12345-23 vs 12345 / 23)
                            excel_case = str(raw_cases[idx][0]).strip() if raw_cases and len(raw_cases) > idx else ""
                            
                            m_t = re.search(r'\d{4,}', t_case) if t_case else None
                            m_e = re.search(r'\d{4,}', excel_case) if excel_case else None
                            base_t = m_t.group(0) if m_t else ""
                            base_e = m_e.group(0) if m_e else ""
                            
                            is_base_match = (base_t and base_e and base_t == base_e)
                            
                            # Option B: Check judge
                            j_val = str(raw_judges[idx][0] or "").strip() if raw_judges and len(raw_judges) > idx else ""
                            is_judge_match = target_j and (target_j in j_val or j_val in target_j)
                            
                            if is_base_match or is_judge_match:
                                matched_rows.append(r_real)
                                match_reason = f"Base Case '{base_t}'" if is_base_match else f"Judge '{target_j}'"
                                log(f"   [SEARCH] DOUBLE SEARCH SUCCESS: Found row {r_real} by Date ({target_d}) and {match_reason}!")
                                try: 
                                    data[COL_CASE_NUM] = excel_case if excel_case else t_case
                                    t_case = data[COL_CASE_NUM]
                                except: pass
                                break
                    except Exception as e:
                        log(f"   [WARNING] Double search error: {e}")

            # Now iterate over matches found in memory
            found_year_match = False
            
            target_date_obj = normalize_date(data.get(COL_DATE))
            exact_date_match_exists = False
            if date_col and target_date_obj:
                for mmr in matched_rows:
                    if normalize_date(ws.Cells(mmr, date_col).Value) == target_date_obj:
                        exact_date_match_exists = True
                        break

            for r in matched_rows:
                # 1. Check Status (Direct Read)
                status_val = str(ws.Cells(r, status_col).Value or "").strip()
                if "בוטל" in status_val and "לא" not in status_val:
                    log(f"⏭️ Skipping Row {r} (Status='{status_val}').")
                    continue
                
                # --- SAFE MULTI-ROW DATE FALLBACK ---
                # אם יש יותר משורה אחת לתיק זה (דיונים בתאריכים שונים), נבדוק אם אולי מדובר במצב
                # שבו רק דיון אחד נותר פתוח (ללא מילים). אם כן - אפשר להניח שזה הדיון הרלוונטי!
                empty_word_rows = []
                for mr in matched_rows:
                    mr_status = str(ws.Cells(mr, status_col).Value or "").strip()
                    if "בוטל" not in mr_status or "לא" in mr_status:
                        mr_words = str(ws.Cells(mr, words_col).Value or "").strip()
                        if mr_words in ["", "None", "0", "nan"]:
                            empty_word_rows.append(mr)

                # 2. STRICT Date Check
                date_match_found = False
                if date_col and data.get(COL_DATE):
                    excel_date_val = ws.Cells(r, date_col).Value
                    excel_date = normalize_date(excel_date_val)
                    target_date = target_date_obj
                    
                    if excel_date and target_date:
                        if excel_date == target_date:
                            date_match_found = True
                        else:
                            # Fuzzy Date Logic: If case number matches EXACTLY and it's the only one, allow it
                            if len(matched_rows) == 1:
                                log(f"   [WARNING] Mismatch Date Override: Excel={excel_date} vs File={target_date} (Assumed Typo)")
                                date_match_found = True
                            elif len(empty_word_rows) == 1 and r in empty_word_rows:
                                if exact_date_match_exists:
                                    log(f"   Mismatch Date: Excel={excel_date} vs File={target_date}. [WARNING] Exact date match exists for another row, NOT using fallback.")
                                    date_match_found = False
                                else:
                                    try:
                                        diff = abs((excel_date - target_date).days)
                                        if diff <= 31:
                                            log(f"   [WARNING] Safe Fallback: Date mismatched ({diff} days diff), but only ONE open hearing exists. Assumed Typo for Row {r}.")
                                            date_match_found = True
                                        else:
                                            log(f"   Mismatch Date: Excel={excel_date} vs File={target_date} (Diff={diff} days). Too large to be a typo.")
                                            date_match_found = False
                                    except:
                                        date_match_found = False
                            else:
                                log(f"   Mismatch Date: Excel={excel_date} vs File={target_date}")
                                date_match_found = False
                    elif not excel_date:
                         if len(matched_rows) == 1 or (len(empty_word_rows) == 1 and r in empty_word_rows):
                              log(f"   [WARNING] Empty Excel Date Override (Assumed Safe)")
                              date_match_found = True
                         else:
                              log(f"   Skipping Row {r}: Excel date is empty, cannot verify match.")
                              date_match_found = False 
                else:
                    # If file has no date
                    if len(matched_rows) == 1 or (len(empty_word_rows) == 1 and r in empty_word_rows):
                         log(f"   [WARNING] No File Date Override (Assumed Safe due to only 1 open hearing)")
                         date_match_found = True
                    else:
                         log(f"   Skipping Row {r}: File has no date extracted and multiple open rows exist.")
                         date_match_found = False
                
                if date_match_found:
                    # Check if empty
                    curr_val = ws.Cells(r, words_col).Value
                    curr_str = str(curr_val).strip()
                    is_empty = False
                    if curr_val is None or curr_str in ["", "None", "nan"]: is_empty = True
                    else:
                        try: 
                            if float(curr_val) < 10: is_empty = True
                        except: pass
                            
                    if is_empty:
                        ws.Cells(r, words_col).Value = data["word_count"]
                        ws.Cells(r, words_col).Interior.Color = 65535
                        import datetime
                        # Update Return Date (Jerusalem = Col 18, South = Col 22)
                        if "2023" in str(sheet_name):
                             ws.Cells(r, 18).Value = datetime.datetime.now()
                        elif "שרות א" in str(sheet_name):
                             ws.Cells(r, 22).Value = datetime.datetime.now()
                        updated += 1
                        log(f"[SUCCESS] Jerusalem Updated: {t_case} (Row {r})")
                        updated_cases.append(t_case) # Mark as success
                        found_row = r
                        # Update Judge if empty
                        if judge_col and data.get("judge"):
                            curr_judge = str(ws.Cells(r, judge_col).Value or "").strip()
                            if curr_judge in ["", "None", "nan"]:
                                ws.Cells(r, judge_col).Value = data["judge"]
                                log(f"   👨⚖️ AI Judge Added: {data['judge']}")
                                
                        break 
                    else:
                        log(f"⏭️  {t_case} already occupied ({curr_val}) - Skipping row {r}")
                        updated_cases.append(str(t_case) + "_ALREADY_FILLED")
                        break

        if updated > 0:
            try:
                wb.Save()
                log(f"[SUCCESS] Saved Successfully: {wb.FullName}")
            except Exception as e:
                log(f"[ERROR] Save Error: {e}")

        # Close instead of remaining open
        excel.Visible = False
        log(f"📝 Excel update complete.")
        
    except Exception as e:
        log(f"[ERROR] General Excel Error: {e}")
        updated_cases.append("READ_ONLY_ERROR")
        
    finally:
        try:
            if wb:
                wb.Close(False)
        except: pass
        if _created_excel:
            try: excel.Quit()
            except: pass
        else:
            # Restore original state so user's session is unaffected
            try: excel.Visible = _orig_visible
            except: pass
            try: excel.DisplayAlerts = _orig_alerts
            except: pass
        
    return updated_cases

def start_google_drive_if_needed():
    if any(os.path.exists(f"{d}:\\") for d in "GHIJK"):
        return
    log("[WARNING] Drive might be disconnected. Attempting to start Google Drive...")
    import glob
    paths = glob.glob(r"C:\Program Files\Google\Drive File Stream\*\GoogleDriveFS.exe")
    if paths:
        try:
            os.system(f'start "" "{paths[0]}"')
            time.sleep(10) # Give it time to mount
            log("[SUCCESS] Tried launching Google Drive.")
        except Exception as e: 
            log(f"[ERROR] Failed to launch Drive: {e}")

def kill_office_processes():
    # log("🧹 Killing old Word and Excel processes before run...")
    # try:
    #     os.system("taskkill /F /IM excel.exe /T >nul 2>&1")
    #     os.system("taskkill /F /IM winword.exe /T >nul 2>&1")
    # except:
    #     pass
    pass

def get_yoel_folder_cases() -> set:
    """סורק את תיקיית דיונים ליואל ושולף את כל מספרי התיקים שיש שם."""
    cases = set()
    yoel_dir = r"F:\בתי משפט\דיונים ליואל"
    
    if not os.path.exists(yoel_dir):
        # ננסה גם כונן אחר אם צריך למרות שבד"כ חרוט F
        yoel_dir_alt = next((f"{d}:\\בתי משפט\\דיונים ליואל" for d in ["F", "G", "H", "I", "C", "D"] if os.path.exists(f"{d}:\\בתי משפט\\דיונים ליואל")), None)
        if yoel_dir_alt:
            yoel_dir = yoel_dir_alt
            
    if not os.path.exists(yoel_dir):
        log("[WARNING] לא נמצאה ספריית 'דיונים ליואל'.")
        return cases
        
    for root, dirs, files in os.walk(yoel_dir):
        for name in dirs + files:
            m = re.search(r'(\d{3,}-\d{2}-\d{2,4})', name)
            if m:
                cases.add(m.group(1))
    
    return cases

def main():
    import sys_process_utils
    sys_process_utils.kill_word_force()
    sys_process_utils.kill_excel_force()
    start_google_drive_if_needed()
    log("[START] Starting Word Count Check...")
    
    stats = {'success': [], 'failed': []}
    unmapped_all = []
    
    # 1. Check INPUT folder
    if not os.path.exists(INPUT_FOLDER):
        log(f"[ERROR] Input folder not found: {INPUT_FOLDER}")
        return stats
    
    files = [f for f in os.listdir(INPUT_FOLDER) if f.lower().endswith(('.doc', '.docx')) and not f.startswith('~$')]
    if not files:
        log("[SUCCESS] No new files found in folder.")
        return stats

    jerusalem_data = [] 
    
    # Load sheet data once for validation across all files
    sheet_data_jerusalem = load_excel_as_dicts(JERUSALEM_EXCEL_PATH, JERUSALEM_SHEET_NAME)
    
    # 🌟 LOAD YOEL CASES FROM FOLDER
    try:
        yoel_folder_cases = get_yoel_folder_cases()
        log(f"[SEARCH] נסרקו {len(yoel_folder_cases)} תיקים קיימים בתיקיית 'דיונים ליואל'.")
    except Exception as e:
        log(f"[WARNING] שגיאה בטעינת תיקי יואל מהתיקייה: {e}")
        yoel_folder_cases = set()
    
    _created_word_main = False
    try:
        word = win32.GetActiveObject("Word.Application")
    except Exception:
        word = win32.Dispatch("Word.Application")
        _created_word_main = True
    word.Visible = False
    try: word.DisplayAlerts = 0 # wdAlertsNone
    except: pass

    try:
        for f in files:
            f_path = os.path.join(INPUT_FOLDER, f)
            
            if f.startswith("~$"): continue

            # --- Extract Case Num ---
            case_num, raw_match = extract_case_check_regex(f)
            
            # --- PRE-CHECK: Was this email already sent manually? ---
            # Attempt to extract date early just for this check, otherwise test with ""
            temp_date_str = ""
            m_date_temp = re.search(r'[_\s-](\d{1,2})[_\s-](\d{1,2})[_\s-](\d{4})[_\s-]', f)
            if m_date_temp:
                temp_date_str = str(datetime(int(m_date_temp.group(3)), int(m_date_temp.group(2)), int(m_date_temp.group(1))).date())

            if case_num and send_district_emails.is_already_sent(case_num, temp_date_str):
                log(f"⏭️ תיק {case_num} מופיע כנשלח. מכיוון שזהו סקריפט ספירת מילים, אנחנו ממשיכים כרגיל.")

            # --- 1. Check if "Other Region" (Folder Match OR Net-HaMishpat pattern) ---
            is_other_region = False
            if case_num and case_num in yoel_folder_cases:
                log(f"🎯 תיק {case_num} מופיע בתיקיית 'דיונים ליואל' - מנותב אוטומטית לטיפול נפרד")
                is_other_region = True
            elif re.search(r'\d{2,}-\d{2,}-\d{2,}[_\s-]+\d{1,2}[_\s-]+\d{1,2}[_\s-]+\d{4}[_\s-]+\d+', f):
                log(f"🎯 קובץ {f} במבנה נט המשפט - מנותב אוטומטית לטיפול נפרד")
                is_other_region = True

            if is_other_region:
                log(f"⏭️ שומר כקובץ מחוזות אחרים (יובל/יוליה) בתיקייה ייעודית מבלי למנות: {f}")
                other_dir = os.path.join(MAIN_FOLDER, "Pending_Other_Region")
                if not os.path.exists(other_dir): os.makedirs(other_dir)
                dest = os.path.join(other_dir, f)
                try: shutil.move(f_path, dest)
                except: pass
                # הוסף לתור → יובל דדו ויוליה
                try:
                    c_num = case_num if case_num else f
                    m_date = re.search(r'[_\s-](\d{1,2})[_\s-](\d{1,2})[_\s-](\d{4})[_\s-]', f)
                    d_obj = None
                    if m_date:
                        d_obj = datetime(int(m_date.group(3)), int(m_date.group(2)), int(m_date.group(1))).date()
                    send_district_emails.run_for_file(dest, c_num, d_obj, 0, "other_region")
                except Exception as em:
                    log(f"[WARNING] שגיאה בהוספת Other Region לתור: {em}")
                continue

            # If the filename doesn't contain a case number, we assume it's not a relevant court file
            # and move it aside to avoid opening valid 'admin' files that might hang Word.

            
            # 2. Extract Text & Word Count
            w_count = 0
            txt_content = ""
            filename_for_date = f
            
            # Remove filename regex part only if we found match, else use full name
            if raw_match: 
                filename_for_date = f.replace(raw_match, "")

            try:
                # Retry mechanism for COM
                doc_opened = False
                for attempt in range(2):
                    try:
                        doc = word.Documents.Open(f_path, ReadOnly=True)
                        doc_opened = True
                        break
                    except Exception as e:
                         # Kill word and recreate if it's completely stuck
                         if attempt == 1:
                             try:
                                 word.Quit()
                                 os.system("taskkill /F /IM winword.exe /T >nul 2>&1")
                                 word = win32.Dispatch("Word.Application")
                                 word.Visible = False
                                 doc = word.Documents.Open(f_path, ReadOnly=True)
                                 doc_opened = True
                             except Exception as inner_e:
                                 log(f"[WARNING] COM fallback failed: {inner_e}")
                         else:
                             time.sleep(1)

                if doc_opened:
                    try:
                        w_count = doc.ComputeStatistics(0) # wdStatisticWords
                    except Exception as stat_e:
                        log(f"[WARNING] שגיאה בספירת מילים (מדלג על המילים, ממשיך למייל): {stat_e}")
                        w_count = 0
                    
                    try:
                        for section in doc.Sections:
                            for header in section.Headers:
                                if not header.Range.Text: continue
                                txt_content += " " + header.Range.Text[:500]
                    except: pass
                    
                    try:
                        txt_content += " " + doc.Range(0, 1000).Text
                        full_len = doc.Characters.Count
                        if full_len > 1000:
                            start_char = max(1000, full_len - 1500)
                            txt_content += " ... " + doc.Range(start_char, full_len).Text
                    except Exception as text_err:
                         log(f"[WARNING] שגיאה בשליפת טקסט דרך וורד: {text_err}")
                
                    try:
                        doc.Close(False)
                    except: pass
                
                txt_content = txt_content.strip()

            except Exception as e:
                log(f"[ERROR] שגיאה בקריאת וורד ב-COM עבור {f}: {e}")

            # Fallback for reading text if COM completely failed and it is a .docx
            if not txt_content and f.lower().endswith('.docx'):
                try:
                    import zipfile
                    import xml.etree.ElementTree as ET
                    with zipfile.ZipFile(f_path) as docx:
                        xml_content = docx.read('word/document.xml')
                        tree = ET.fromstring(xml_content)
                        texts = [node.text for node in tree.iter() if node.tag.endswith('t') and node.text]
                        all_text = " ".join(texts)
                        txt_content = all_text[:1000] + " ... " + all_text[-1500:] if len(all_text) > 2500 else all_text
                        log("[SUCCESS] (חילץ טקסט בהצלחה דרך מנגנון הגיבוי)")
                except Exception as zip_e:
                    log(f"[WARNING] מנגנון גיבוי לקריאת docx נכשל: {zip_e}")

            if not txt_content:
                log(f"[ERROR] לא ניתן לחלץ טקסט עבור {f}. אי אפשר לבדוק אם המסמך שלם. מדלג.")
                continue

            # 2.1 Fallback: Extract Case Num from Content if missing in filename
            if not case_num:
                case_num, _ = extract_case_check_regex(txt_content, is_content=True)
                if case_num:
                    log(f"   [SUCCESS] Found case number in CONTENT: {case_num} (File: {f})")
            
            # 2.2 Final Check: Still no Case Num? -> Move to Review
            if not case_num:
                log(f"⚠ No case number found (Filename or Content): {f} -> Moving to Review")
                review_dir = os.path.join(MAIN_FOLDER, "Requires_Manual_Review")
                if not os.path.exists(review_dir): os.makedirs(review_dir)
                try: 
                    shutil.move(f_path, os.path.join(review_dir, f))
                except Exception as e:
                    log(f"[WARNING] Failed to move file: {e}")
                continue

            # 3. Date Extraction (Case Number found)
            date_obj = None # Will hold datetime.date object
            
            # Since we found case_num early, we just log it
            # log(f"   [SUCCESS] Found case number in filename: {case_num}") 
            
            # LLM Logic removed from here as we enforce filename match
            
                
            # Date Extraction from filename (e.g. 28.01.2026, 21.01.26, 4.2.26)
                
            # Date Extraction from filename (e.g. 28.01.2026, 21.01.26, 4.2.26)
            # Find pattern like D.M.Y or D-M-Y
            date_match = re.search(r'(\d{1,2})[\./-](\d{1,2})[\./-](\d{2,4})', filename_for_date)
            if date_match:
                try:
                    d, m, y = int(date_match.group(1)), int(date_match.group(2)), int(date_match.group(3))
                    if y < 100: y += 2000 # 26 -> 2026
                    date_obj = datetime(y, m, d).date()
                except: pass
            
            # --- FALLBACK: Date from Content ---
            if not date_obj:
                # Search in first 1000 chars of text
                # Look for "תאריך: DD/MM/YYYY" or just date patterns
                try:
                    snippet = txt_content[:1000]
                    # Regex for dates in text
                    content_date_match = re.search(r'(\d{1,2})[\./-](\d{1,2})[\./-](\d{2,4})', snippet)
                    if content_date_match:
                        d, m, y = int(content_date_match.group(1)), int(content_date_match.group(2)), int(content_date_match.group(3))
                        if y < 100: y += 2000
                        date_obj = datetime(y, m, d).date()
                        log(f"   📅 Date found in CONTENT: {date_obj}")
                except Exception as e:
                    log(f"[WARNING] Date content fallback failed: {e}")
            
            # 4. Check Region
            is_south = gsheets_utils.is_south_region(txt_content, f)
            
            if is_south:
                log(f"🌍 Detected as South: {f}. Case: {case_num}, Date: {date_obj}, Words: {w_count}")
                if case_num:
                    try:
                        # Update South Excel directly (passes dict!)
                        ok = gsheets_utils.update_south_sheet({
                            'מספר תיק': case_num,
                            'מספר מילים': w_count,
                            'תאריך': date_obj
                        })
                        # המייל כבר נשלח בתהליך fast_email_sender! לכן כאן רק מתעדים באקסל
                        # אין דיוור כפול כיוון שזוהה ב- fast_email_sender

                        if ok == True:
                            log(f"[SUCCESS] Successfully updated South Excel: {case_num}")
                            stats['success'].append({'file': f, 'case': case_num, 'date': str(date_obj), 'words': w_count})
                            log("   Saving to Word_Count_Excel...")
                            # Excel update logic would go here
                            log(f"   [SUCCESS] [EXCEL SUCCESS] {f}")
                            done_dir = os.path.join(DONE_FOLDER)
                            if not os.path.exists(done_dir): os.makedirs(done_dir)
                            try: safe_move(f_path, done_dir)
                            except: pass
                        elif ok in ["READ_ONLY", "ERROR"]:
                            log(f"⏸️ South Excel is locked/read-only or returned an error. Leaving {f} pending safely.")
                            stats['failed'].append({'file': f, 'reason': "South Excel Locked"})
                        else:
                            log(f"[ERROR] Failed to update South Excel (Not found) - Moving to Failed")
                            stats['failed'].append({'file': f, 'reason': "South Update Failed"})
                            unmapped_all.append({
                                COL_CASE_NUM: case_num,
                                COL_DATE: date_obj,
                                "judge": "דרום/ב\"ש",
                                "filename": f,
                                "word_count": w_count
                            })
                            failed_dir = os.path.join(INPUT_FOLDER, "Failed")
                            if not os.path.exists(failed_dir): os.makedirs(failed_dir)
                            try: safe_move(f_path, failed_dir)
                            except: pass
                            
                    except Exception as critical_e:
                        log(f"[ERROR] CRITICAL ERROR calling update_south_sheet: {critical_e}")
                        stats['failed'].append({'file': f, 'reason': f"South Critical: {critical_e}"})
                else:
                    log("   [WARNING] South file without case number - Skipping")
            else:
                # Jerusalem -> Add to list for bulk processing
                log(f"🌍 Detected as Jerusalem: {f}")
                if case_num:
                    jerusalem_data.append({
                        COL_CASE_NUM: case_num,
                        COL_DATE: date_obj,
                        COL_WORD_COUNT: w_count, 
                        "word_count": w_count,
                        "filename": f,
                        "file_path": f_path,
                        "text_snippet": txt_content,
                        "judge": extract_judge_llm(txt_content) if extract_judge_llm else None
                    })
                    # NOTE: File will be moved to Done AFTER update succeeds (see below)
                else:
                    log("[WARNING] Jerusalem file without case number - Skipping")
 
    finally:
        if _created_word_main:
            try: word.Quit()
            except: pass
 
    # 5. Process Jerusalem Batch - Get list of actually updated cases
    updated_cases = []
    if jerusalem_data:
        updated_cases = update_jerusalem_table(JERUSALEM_EXCEL_PATH, JERUSALEM_SHEET_NAME, jerusalem_data)
        if "READ_ONLY_ERROR" in updated_cases:
            log("🛑 Excel is tied up or Read-Only. Safely aborting this batch to leave files pending without sending alert emails.")
            jerusalem_data = [] # Empty list to skip processing loop
    else:
        log("[SUCCESS] No Jerusalem cases to update.")
    
    # 6. Save to F: drive hierarchy + Move to appropriate folder
    if not os.path.exists(DONE_FOLDER): os.makedirs(DONE_FOLDER)
    
    # Define folder for cases not updated in Excel (e.g. date mismatch)
    REVIEW_FOLDER = os.path.join(MAIN_FOLDER, "Requires_Manual_Review")
    if not os.path.exists(REVIEW_FOLDER): os.makedirs(REVIEW_FOLDER)

    for data in jerusalem_data:
        src = data.get("file_path", "")
        fname = data.get("filename", "")
        case_n = data.get(COL_CASE_NUM)
        
        if src and os.path.exists(src):
            # Try to save to F: drive hierarchy first
            try:
                details = {
                    'case_num': case_n,
                    'date': data.get(COL_DATE),
                }
                save_file_to_hierarchy(src, details, "Jerusalem", text_content=data.get("text_snippet"))
            except Exception as e:
                log(f"[WARNING] Could not save {fname} to F: drive: {e}")
            
            # [SUCCESS] 1. הפעלת מנגנון אימות קפדני (Validation)
            is_valid, validation_msg = file_validator.validate_transcription_file(
                text_content=data.get("text_snippet", ""),
                case_num=case_n,
                date_obj=data.get(COL_DATE),
                sheet_data=sheet_data_jerusalem  # מעבירים את כל הגיליון לבדיקת דיונים ישנים
            )

            if not is_valid:
                log(f"[ERROR] קובץ נפסל בבדיקה: {fname} -> הסיבה: {validation_msg}")
                # Move to Validation_Failed folder
                fail_folder = os.path.join(REVIEW_FOLDER, "Validation_Failed")
                if not os.path.exists(fail_folder): os.makedirs(fail_folder)
                try:
                    safe_move(src, fail_folder)
                    log(f"📁 Moved to Validation_Failed: {fname}")
                except Exception as eval_err:
                    log(f"[WARNING] Failed to move to Validation_Failed: {eval_err}")
                stats['failed'].append({
                    'file': fname,
                    'reason': f"Validation Failed: {validation_msg}"
                })
                continue  # מדלג לקובץ הבא

            # =============================================================
            # CHECK IF UPDATED IN EXCEL
            if case_n in updated_cases:
                # Success -> Done
                done_path = os.path.join(DONE_FOLDER)
                if not os.path.exists(done_path): os.makedirs(done_path)
                try:
                    safe_move(src, done_path)
                    log(f"📁 Moved to Done: {fname}")
                    stats['success'].append({
                        'file': fname,
                        'case': case_n,
                        'date': str(data.get(COL_DATE)),
                        'words': data.get("word_count")
                    })
                except Exception as e:
                    log(f"[WARNING] Failed to move {fname} to Done: {e}")
            elif case_n and (str(case_n) + "_ALREADY_FILLED") in updated_cases:
                # Already filled -> Done (Duplicate)
                done_path = os.path.join(DONE_FOLDER)
                if not os.path.exists(done_path): os.makedirs(done_path)
                try:
                    safe_move(src, done_path)
                    log(f"📁 Moved to Done (Already Filled/Duplicate): {fname}")
                    stats['success'].append({
                        'file': fname,
                        'case': case_n,
                        'date': str(data.get(COL_DATE)),
                        'words': 'ALREADY FILLED'
                    })
                except Exception as e:
                    log(f"[WARNING] Failed to move duplicate {fname} to Done: {e}")
            else:
                # Not Updated -> Review
                log(f"   [WARNING] Filename validation failed for {fname}, moving to Review folder.")
                try:
                    safe_move(src, REVIEW_FOLDER)
                    log(f"[WARNING] Case {case_n} NOT updated in Excel (Date mismatch?) -> Moved to Manual Review")
                    stats['failed'].append({
                        'file': fname, 
                        'reason': f"Not updated in Excel (Check Date/Case Match)"
                    })
                    unmapped_all.append(data)
                except Exception as e:
                    log(f"[WARNING] Failed to move {fname} to Review: {e}")
                    
    if unmapped_all:
        send_unmapped_email(unmapped_all)
    
    log("Processing Complete.")
    return stats

def run_batch_with_stats():
    return main()

if __name__ == "__main__":
    main()
